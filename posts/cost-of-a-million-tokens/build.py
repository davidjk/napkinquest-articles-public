"""
AI True Cost Model — Consolidated Build Script
================================================

Builds the complete AI True Cost Model spreadsheet from scratch.

Produces an .xlsx file with thirteen sheets:
   1. README                       — methodology, color conventions, caveats
   2. Assumptions                  — all editable inputs (blue cells)
   3. Cost Build                   — five-layer cost decomposition per million tokens
   4. Revenue Build                — consumer subs, API, enterprise, government revenue
   5. Unit Economics               — cost vs revenue at multiple levels
   6. Sensitivity                  — single-variable sensitivity tables
   7. Build vs Rent                — capex ownership toggle (0% / 25% / 50% / 75% / 100%)
   8. Breakeven                    — revenue and scale required to clear cost
   9. No-Subsidy Pricing           — consumer sub cost per usage tier at full-economic
  10. Implicit Demand              — tokens implied by committed multi-year spend
  11. Plausibility                 — three lenses on whether implied demand is real
  12. Capex Perspective            — capex vs US GDP, business investment, grid capacity
  13. Stacks - Pre Announcement    — three-way competitor comparison (legacy baseline)
  14. Stacks - Post Announcement   — same comparison after Apr/May 2026 Anthropic deals

Color conventions (consistent across all sheets):
  Blue text        — hardcoded input values (editable)
  Black text       — formulas and intra-sheet calculations
  Green text       — cross-sheet references
  Yellow fill      — key assumptions to scrutinize
  Yellow-orange    — primary totals
  Light orange     — full-economic-cost totals (Build vs Rent)

Usage:
    python build.py [output_path]

If no output_path is provided, writes to ./model.xlsx in the current directory.

Dependencies:
    openpyxl >= 3.0
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# STYLE CONSTANTS
# ============================================================================

BLUE = Font(name='Arial', color='0000FF', size=10)
BLACK = Font(name='Arial', color='000000', size=10)
BLACK_BOLD = Font(name='Arial', color='000000', size=10, bold=True)
GREEN = Font(name='Arial', color='008000', size=10)
WHITE_BOLD = Font(name='Arial', color='FFFFFF', size=11, bold=True)
TITLE = Font(name='Arial', size=14, bold=True)
SECTION = Font(name='Arial', size=11, bold=True)
NOTE = Font(name='Arial', italic=True, color='595959', size=9)
WARNING = Font(name='Arial', size=10, italic=True, color='C00000', bold=True)

YELLOW = PatternFill('solid', start_color='FFFF00')
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
SECTION_FILL = PatternFill('solid', start_color='D9E1F2')
TOTAL_FILL = PatternFill('solid', start_color='FFF2CC')
PRIMARY_TOTAL = PatternFill('solid', start_color='FFD966')
ECONOMIC_TOTAL = PatternFill('solid', start_color='F4B084')
WARNING_FILL = PatternFill('solid', start_color='FFF0F0')
GOOGLE_FILL = PatternFill('solid', start_color='E8F0FE')
OAI_FILL = PatternFill('solid', start_color='F0E8E8')
ANTH_FILL = PatternFill('solid', start_color='E8F0E8')

thin = Side(border_style='thin', color='BFBFBF')
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(wrap_text=True, vertical='center')


# ============================================================================
# HELPERS
# ============================================================================

def write_header_row(ws, row, headers, fill=HEADER_FILL, font=WHITE_BOLD):
    """Write a styled header row across the columns."""
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = font
        c.fill = fill
        c.alignment = CENTER
        c.border = BORDER


def write_section_label(ws, row, title, span=5):
    """Write a section divider row spanning columns 1..span."""
    c = ws.cell(row=row, column=1, value=title)
    c.font = SECTION
    c.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c.alignment = LEFT


def write_takeaway_block(ws, start_row, takeaways, span=5, row_height=32):
    """Write a list of takeaway bullets, each merged across `span` columns."""
    r = start_row
    for t in takeaways:
        cell = ws.cell(row=r, column=1, value=f"\u2022 {t}")
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        ws.row_dimensions[r].height = row_height
        r += 1
    return r


# ============================================================================
# README SHEET
# ============================================================================

def build_readme(wb):
    ws = wb.active
    ws.title = "README"

    ws['A1'] = "AI True Cost Model"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:D1')

    ws['A3'] = "Purpose"
    ws['A3'].font = SECTION
    ws['A4'] = ("Estimates the all-in cost per million output tokens for a frontier AI model, "
                "broken into FIVE cost layers including capex, and compares against typical "
                "subscription and contract revenue. Also includes a three-way competitive "
                "comparison (Google / OpenAI / Anthropic) before and after the April/May 2026 "
                "Anthropic compute deals.")
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A4:D6')

    ws['A8'] = "Sheets"
    ws['A8'].font = SECTION
    sheets_info = [
        ("1. Assumptions", "All inputs in blue. Edit these to flex the model. Three scenarios: Low / Mid / High."),
        ("2. Cost Build", "Calculates cost per million output tokens across five layers including capex."),
        ("3. Revenue Build", "Models blended revenue from consumer subs, API, enterprise, and government."),
        ("4. Unit Economics", "Compares cost vs revenue per query, per million tokens, and company-wide."),
        ("5. Sensitivity", "Single-variable sensitivity of cost per M tokens to key drivers."),
        ("6. Build vs Rent", "Toggles capex ownership mix between owned datacenter and rented hyperscaler capacity."),
        ("7. Breakeven", "Inverts cost into revenue and scale required to clear the cost stack."),
        ("8. No-Subsidy Pricing", "Consumer subscription cost at full-economic per usage tier."),
        ("9. Implicit Demand", "Token volume implied by committed multi-year compute spend per lab."),
        ("10. Plausibility", "Three lenses on whether the implied demand can be absorbed."),
        ("11. Capex Perspective", "Capex framed against US GDP, business investment, and grid capacity."),
        ("12. Stacks - Pre Announcement", "Three-way competitor comparison before the April/May 2026 deals."),
        ("13. Stacks - Post Announcement", "Updated comparison after Anthropic's Google/xAI/Amazon expansions."),
    ]
    for i, (s, d) in enumerate(sheets_info, start=9):
        ws[f'A{i}'] = s
        ws[f'A{i}'].font = BLACK_BOLD
        ws[f'B{i}'] = d
        ws.merge_cells(f'B{i}:D{i}')

    next_r = 9 + len(sheets_info) + 1
    ws[f'A{next_r}'] = "Color Convention"
    ws[f'A{next_r}'].font = SECTION
    conv = [
        ("Blue", "Hardcoded inputs you can change", BLUE),
        ("Black", "Formulas and calculations", BLACK),
        ("Green", "Cross-sheet references", GREEN),
        ("Yellow fill", "Key assumptions to review", None),
    ]
    conv_start = next_r + 1
    for i, (c, d, f) in enumerate(conv, start=conv_start):
        ws[f'A{i}'] = c
        if f:
            ws[f'A{i}'].font = f
        else:
            ws[f'A{i}'].fill = YELLOW
        ws[f'B{i}'] = d
        ws.merge_cells(f'B{i}:D{i}')

    caveat_label = conv_start + len(conv) + 1
    ws[f'A{caveat_label}'] = "Caveats"
    ws[f'A{caveat_label}'].font = SECTION
    caveats = [
        "Frontier lab cost structures are not public. All inputs are triangulated estimates as of early 2026.",
        "Capex commitments often live in hyperscaler partners' balance sheets via take-or-pay contracts. The Build vs Rent toggle approximates how this allocates.",
        "Inference cost per token has fallen roughly 10x year over year for comparable capability. Treat training amortization as the most volatile input.",
        "Enterprise contract economics vary widely between deals.",
        "The competitor stacks sheets reflect a specific view of relative cost positions; multipliers are editable.",
        "Implicit Demand and Plausibility sheets translate dollars committed into tokens served at current cost - they are scenarios, not forecasts.",
    ]
    for i, c in enumerate(caveats, start=caveat_label + 1):
        ws[f'A{i}'] = f"\u2022 {c}"
        ws[f'A{i}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{i}:D{i}')
        ws.row_dimensions[i].height = 30

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30


# ============================================================================
# ASSUMPTIONS SHEET
# ============================================================================

def build_assumptions(wb):
    """
    Build the Assumptions sheet. Returns a dict mapping logical input keys
    to their row numbers, used by all other sheets for references.
    """
    ws = wb.create_sheet("Assumptions")

    ws['A1'] = "Assumptions"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    write_header_row(ws, 3, ['Driver', 'Unit', 'Low', 'Mid', 'High'])

    rows = {}

    def section_label(r, title):
        write_section_label(ws, r, title)
        return r + 1

    def input_row(r, key, label, unit, low, mid, high, fmt=None, hi=False):
        ws.cell(row=r, column=1, value=label).font = BLACK
        ws.cell(row=r, column=2, value=unit).font = BLACK
        ws.cell(row=r, column=2).alignment = CENTER
        for col, v in zip([3, 4, 5], [low, mid, high]):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = BLUE
            cell.alignment = RIGHT
            if fmt:
                cell.number_format = fmt
            if hi:
                cell.fill = YELLOW
        rows[key] = r

    r = 4
    r = section_label(r, "Layer 1: Training Amortization")
    input_row(r, 'training_cost', "Frontier training run cost", "$M", 100, 500, 1500, '$#,##0;($#,##0);-', True); r += 1
    input_row(r, 'life', "Useful life of model", "months", 24, 18, 12, '0', True); r += 1
    input_row(r, 'industry_tokens', "Trailing 12mo output tokens served (industry)", "trillions", 500, 250, 100, '#,##0', True); r += 1
    input_row(r, 'lab_share', "Lab share of those tokens", "%", 0.30, 0.20, 0.10, '0.0%'); r += 1

    r += 1
    r = section_label(r, "Layer 2: Inference Compute")
    input_row(r, 'gpu_hr_rented', "GPU cost per hour - RENTED (hyperscaler markup)", "$/hr", 1.5, 2.5, 4.0, '$#,##0.00'); r += 1
    input_row(r, 'gpu_hr_owned', "GPU cost per hour - OWNED (operating only, no capex)", "$/hr", 0.6, 1.0, 1.6, '$#,##0.00'); r += 1
    input_row(r, 'tokens_per_hr', "Output tokens per GPU-hour", "tokens/hr", 1500000, 800000, 400000, '#,##0', True); r += 1
    input_row(r, 'io_ratio', "Input:Output token ratio", "ratio", 5, 8, 12, '0.0'); r += 1
    input_row(r, 'input_factor', "Input token cost factor vs output", "%", 0.20, 0.25, 0.30, '0.0%'); r += 1

    r += 1
    r = section_label(r, "Layer 3: Infrastructure Overhead (operating)")
    input_row(r, 'power_pct', "Power as % of compute cost", "%", 0.15, 0.25, 0.40, '0.0%', True); r += 1
    input_row(r, 'net_pct', "Networking, cooling, redundancy", "%", 0.08, 0.12, 0.18, '0.0%'); r += 1
    input_row(r, 'dc_pct', "Datacenter real estate (lease)", "%", 0.03, 0.05, 0.08, '0.0%'); r += 1

    r += 1
    r = section_label(r, "Layer 4: R&D, People, Data")
    input_row(r, 'rd_spend', "Annual R&D spend", "$M", 800, 1500, 3000, '$#,##0', True); r += 1
    input_row(r, 'data_spend', "Annual data acquisition and licensing", "$M", 50, 200, 500, '$#,##0'); r += 1
    input_row(r, 'rlhf', "RLHF and human feedback labor", "$M", 30, 80, 200, '$#,##0'); r += 1

    r += 1
    r = section_label(r, "Layer 5: Capex (amortized)")
    input_row(r, 'dc_capacity_mw', "Datacenter capacity required", "MW", 300, 1000, 3000, '#,##0', True); r += 1
    input_row(r, 'dc_cost_per_mw', "Datacenter buildout cost", "$M/MW", 25, 40, 60, '$#,##0', True); r += 1
    input_row(r, 'dc_useful_life', "Datacenter useful life", "years", 20, 15, 10, '0'); r += 1
    input_row(r, 'gpu_count', "GPU count required at scale", "thousands", 200, 600, 1500, '#,##0', True); r += 1
    input_row(r, 'gpu_unit_cost', "Avg GPU system cost (all-in)", "$K", 30, 50, 80, '$#,##0', True); r += 1
    input_row(r, 'gpu_useful_life', "GPU useful life before refresh", "years", 5, 4, 3, '0', True); r += 1
    input_row(r, 'power_infra_capex', "Power infra capex (substations, etc)", "$M", 500, 2000, 6000, '$#,##0'); r += 1
    input_row(r, 'power_infra_life', "Power infra useful life", "years", 25, 20, 15, '0'); r += 1
    input_row(r, 'capex_ownership_pct', "% of capex owned vs rented (0=all rent, 100=all own)", "%", 0.0, 0.5, 1.0, '0.0%', True); r += 1

    r += 1
    r = section_label(r, "Revenue Inputs")
    input_row(r, 'subs', "Consumer paid subscribers", "millions", 5, 15, 30, '#,##0.0'); r += 1
    input_row(r, 'sub_price', "Avg consumer subscription price", "$/mo", 18, 20, 22, '$#,##0.00'); r += 1
    input_row(r, 'queries_pm', "Avg consumer queries per user per month", "queries", 100, 250, 500, '#,##0'); r += 1
    input_row(r, 'tokens_per_q', "Avg output tokens per consumer query", "tokens", 400, 600, 1000, '#,##0'); r += 1
    input_row(r, 'api_rev', "API revenue (annual)", "$M", 200, 600, 1500, '$#,##0', True); r += 1
    input_row(r, 'ent_rev', "Enterprise contract revenue (annual)", "$M", 500, 2000, 5000, '$#,##0', True); r += 1
    input_row(r, 'gov_rev', "Government contract revenue (annual)", "$M", 50, 250, 800, '$#,##0'); r += 1

    r += 1
    r = section_label(r, "Demand and Breakeven Inputs")
    input_row(r, 'light_q', "Light user queries per month", "queries", 50, 50, 50, '#,##0'); r += 1
    input_row(r, 'light_tpq', "Light user output tokens per query", "tokens", 300, 400, 600, '#,##0'); r += 1
    input_row(r, 'power_q', "Power user queries per month", "queries", 1000, 2000, 4000, '#,##0'); r += 1
    input_row(r, 'power_tpq', "Power user output tokens per query", "tokens", 600, 1000, 1500, '#,##0'); r += 1
    input_row(r, 'anthropic_5y_b', "Anthropic 5-year compute commitment", "$B", 150, 200, 250, '$#,##0', True); r += 1
    input_row(r, 'openai_5y_b', "OpenAI 5-year compute commitment", "$B", 200, 300, 400, '$#,##0', True); r += 1
    input_row(r, 'baseline_lab_5y_b', "Baseline lab 5-year commitment (xAI/Google/Meta avg)", "$B", 50, 100, 200, '$#,##0', True); r += 1
    input_row(r, 'commitment_years', "Commitment period", "years", 5, 5, 5, '0'); r += 1
    input_row(r, 'heavy_user_tokens_yr', "Heavy user annual output tokens", "M tokens/yr", 12, 24, 60, '#,##0.0'); r += 1
    input_row(r, 'seat_price', "Enterprise seat price (per seat per year)", "$/seat/yr", 1200, 3000, 6000, '$#,##0'); r += 1

    r += 1
    r = section_label(r, "Macro Benchmark Inputs")
    input_row(r, 'us_gdp_t', "US nominal GDP", "$T", 28, 29, 30, '$#,##0.0'); r += 1
    input_row(r, 'us_business_capex_t', "US nonresidential fixed investment (annual)", "$T/yr", 3.3, 3.5, 3.8, '$#,##0.0'); r += 1
    input_row(r, 'us_grid_peak_gw', "US grid peak demand", "GW", 700, 750, 800, '#,##0'); r += 1
    input_row(r, 'us_households_m', "US households", "millions", 130, 131, 132, '#,##0'); r += 1
    input_row(r, 'capex_share_of_spend', "Hardware/datacenter share of committed spend", "%", 0.50, 0.60, 0.70, '0.0%'); r += 1

    ws.column_dimensions['A'].width = 56
    ws.column_dimensions['B'].width = 14
    for c in ['C', 'D', 'E']:
        ws.column_dimensions[c].width = 14

    return rows


def assumption_ref(rows, key, col):
    """Build a cross-sheet reference to an Assumptions row."""
    return f"Assumptions!{col}{rows[key]}"


# ============================================================================
# COST BUILD SHEET
# ============================================================================

def build_cost_build(wb, rows):
    """
    Build the Cost Build sheet. Returns a dict of row numbers for key
    output rows, used by downstream sheets.
    """
    ws = wb.create_sheet("Cost Build")

    ws['A1'] = "Cost Per Million Output Tokens"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    ws['A2'] = "Includes Layer 5 capex. Build-vs-rent mix is set in Assumptions (capex_ownership_pct)."
    ws['A2'].font = NOTE
    ws.merge_cells('A2:E2')

    write_header_row(ws, 4, ['Component', 'Formula Logic', 'Low', 'Mid', 'High'])

    cb = {}

    def aref(key, col):
        return assumption_ref(rows, key, col)

    # ---- Layer 1: Training ----
    r = 5
    write_section_label(ws, r, "Layer 1: Training amortization"); r += 1

    ws.cell(row=r, column=1, value="Training cost ($M)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('training_cost', scen)}")
        cell.font = GREEN
        cell.number_format = '$#,##0;($#,##0);-'
    cb['training_cost'] = r; r += 1

    ws.cell(row=r, column=1, value="Useful life (months)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('life', scen)}")
        cell.font = GREEN
        cell.number_format = '0'
    cb['life'] = r; r += 1

    ws.cell(row=r, column=1, value="Lab tokens annually (trillions)").font = BLACK
    ws.cell(row=r, column=2, value="Industry tokens x lab share").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col,
                       value=f"={aref('industry_tokens', scen)}*{aref('lab_share', scen)}")
        cell.font = BLACK
        cell.number_format = '#,##0.0'
    cb['lab_tokens'] = r; r += 1

    ws.cell(row=r, column=1, value="$/M output tokens (training)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="Training$ / (lab tokens x life/12)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['training_cost']}/({cl}{cb['lab_tokens']}*{cl}{cb['life']}/12)")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00;($#,##0.00);-'
        cell.fill = TOTAL_FILL
    cb['training_per_m'] = r; r += 2

    # ---- Layer 2: Inference ----
    write_section_label(ws, r, "Layer 2: Inference compute"); r += 1

    ws.cell(row=r, column=1, value="Effective GPU $/hr (blended own/rent)").font = BLACK
    ws.cell(row=r, column=2, value="own_pct x owned_rate + (1-own_pct) x rented_rate").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col,
                       value=f"={aref('capex_ownership_pct', scen)}*{aref('gpu_hr_owned', scen)}+"
                             f"(1-{aref('capex_ownership_pct', scen)})*{aref('gpu_hr_rented', scen)}")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    cb['gpu_hr'] = r; r += 1

    ws.cell(row=r, column=1, value="Output tokens per GPU-hr").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('tokens_per_hr', scen)}")
        cell.font = GREEN
        cell.number_format = '#,##0'
    cb['tokens_per_hr'] = r; r += 1

    ws.cell(row=r, column=1, value="Output compute $/M tokens").font = BLACK
    ws.cell(row=r, column=2, value="GPU $/hr x 1M / tokens per hr").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['gpu_hr']}*1000000/{cl}{cb['tokens_per_hr']}")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    cb['output_compute'] = r; r += 1

    ws.cell(row=r, column=1, value="Input compute $/M tokens").font = BLACK
    ws.cell(row=r, column=2, value="Output compute x ratio x cost factor").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['output_compute']}*{aref('io_ratio', scen)}*{aref('input_factor', scen)}")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    cb['input_compute'] = r; r += 1

    ws.cell(row=r, column=1, value="$/M output tokens (inference)").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['output_compute']}+{cl}{cb['input_compute']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00'
        cell.fill = TOTAL_FILL
    cb['inference_per_m'] = r; r += 2

    # ---- Layer 3: Infra overhead ----
    write_section_label(ws, r, "Layer 3: Infrastructure overhead (operating)"); r += 1

    for label, key in [("Power overhead", 'power_pct'),
                        ("Networking, cooling, redundancy", 'net_pct'),
                        ("Datacenter real estate (lease)", 'dc_pct')]:
        ws.cell(row=r, column=1, value=label).font = BLACK
        for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col, value=f"={cl}{cb['inference_per_m']}*{aref(key, scen)}")
            cell.font = BLACK
            cell.number_format = '$#,##0.00'
        cb[key.replace('_pct', '')] = r
        r += 1

    ws.cell(row=r, column=1, value="$/M output tokens (infra)").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['power']}+{cl}{cb['net']}+{cl}{cb['dc']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00'
        cell.fill = TOTAL_FILL
    cb['infra_per_m'] = r; r += 2

    # ---- Layer 4: R&D ----
    write_section_label(ws, r, "Layer 4: R&D, people, data"); r += 1

    ws.cell(row=r, column=1, value="Annual R&D + data + RLHF ($M)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col,
                       value=f"={aref('rd_spend', scen)}+{aref('data_spend', scen)}+{aref('rlhf', scen)}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    cb['rd_total'] = r; r += 1

    ws.cell(row=r, column=1, value="$/M output tokens (R&D)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="R&D$ / annual lab tokens").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{cb['rd_total']}/{cl}{cb['lab_tokens']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00'
        cell.fill = TOTAL_FILL
    cb['rd_per_m'] = r; r += 2

    # ---- Layer 5: Capex ----
    write_section_label(ws, r, "Layer 5: Capex amortization"); r += 1

    ws.cell(row=r, column=1, value="Datacenter capex total ($M)").font = BLACK
    ws.cell(row=r, column=2, value="capacity_MW x cost_per_MW").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col,
                       value=f"={aref('dc_capacity_mw', scen)}*{aref('dc_cost_per_mw', scen)}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    cb['dc_capex'] = r; r += 1

    ws.cell(row=r, column=1, value="GPU capex total ($M)").font = BLACK
    ws.cell(row=r, column=2, value="gpu_count(K) x unit_cost(K)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        # gpu_count in thousands × unit_cost in $K = total in $M (since K×K=M)
        cell = ws.cell(row=r, column=col,
                       value=f"={aref('gpu_count', scen)}*{aref('gpu_unit_cost', scen)}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    cb['gpu_capex'] = r; r += 1

    ws.cell(row=r, column=1, value="Power infra capex ($M)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('power_infra_capex', scen)}")
        cell.font = GREEN
        cell.number_format = '$#,##0'
    cb['power_capex'] = r; r += 1

    ws.cell(row=r, column=1, value="Annual depreciation ($M)").font = BLACK
    ws.cell(row=r, column=2, value="DC/life + GPU/life + power_infra/life").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['dc_capex']}/{aref('dc_useful_life', scen)}+"
                             f"{cl}{cb['gpu_capex']}/{aref('gpu_useful_life', scen)}+"
                             f"{cl}{cb['power_capex']}/{aref('power_infra_life', scen)}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    cb['annual_depr'] = r; r += 1

    ws.cell(row=r, column=1, value="Capex burden borne by lab (own pct)").font = BLACK
    ws.cell(row=r, column=2, value="Annual depr x ownership %").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['annual_depr']}*{aref('capex_ownership_pct', scen)}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    cb['capex_borne'] = r; r += 1

    ws.cell(row=r, column=1, value="$/M output tokens (capex - lab)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="Capex borne / lab annual tokens").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{cb['capex_borne']}/{cl}{cb['lab_tokens']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00'
        cell.fill = TOTAL_FILL
    cb['capex_per_m'] = r; r += 1

    ws.cell(row=r, column=1, value="$/M output tokens (capex - in GPU rate, rented portion)").font = BLACK
    ws.cell(row=r, column=2, value="Already inside Layer 2 via rented GPU markup").font = NOTE
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col,
                       value=f"=({aref('gpu_hr_rented', scen)}-{aref('gpu_hr_owned', scen)})*"
                             f"1000000/{aref('tokens_per_hr', scen)}*"
                             f"(1-{aref('capex_ownership_pct', scen)})*"
                             f"(1+{aref('io_ratio', scen)}*{aref('input_factor', scen)})")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    cb['capex_via_rent'] = r; r += 1

    ws.cell(row=r, column=1, value="$/M output tokens (capex TOTAL economic)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="Lab capex burden + rent markup (full economic cost)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['capex_per_m']}+{cl}{cb['capex_via_rent']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00'
        cell.fill = TOTAL_FILL
    cb['capex_total_per_m'] = r; r += 2

    # ---- Total lab cash cost ----
    ws.cell(row=r, column=1, value="TOTAL TRUE COST per M output tokens").font = Font(name='Arial', size=12, bold=True)
    ws.cell(row=r, column=1).fill = PRIMARY_TOTAL
    ws.cell(row=r, column=2, value="Sum of layers 1-4 + capex borne by lab only").font = NOTE
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{cb['training_per_m']}+{cl}{cb['inference_per_m']}+"
                             f"{cl}{cb['infra_per_m']}+{cl}{cb['rd_per_m']}+{cl}{cb['capex_per_m']}")
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.number_format = '$#,##0.00'
        cell.fill = PRIMARY_TOTAL
    cb['total'] = r; r += 1

    # ---- Total economic cost (includes hyperscaler-borne capex) ----
    ws.cell(row=r, column=1, value="TOTAL ECONOMIC COST per M output tokens").font = Font(name='Arial', size=12, bold=True)
    ws.cell(row=r, column=1).fill = ECONOMIC_TOTAL
    ws.cell(row=r, column=2, value="Includes capex borne by hyperscaler partners too").font = NOTE
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        own_rate = aref('gpu_hr_owned', scen)
        tph = aref('tokens_per_hr', scen)
        io_r = aref('io_ratio', scen)
        in_f = aref('input_factor', scen)
        pp = aref('power_pct', scen)
        np_ = aref('net_pct', scen)
        dp = aref('dc_pct', scen)
        # Build inference at owned operating rate, then layer infra on top, then full depreciation
        inf_owned = f'({own_rate}*1000000/{tph}*(1+{io_r}*{in_f}))'
        infra_owned = f'{inf_owned}*({pp}+{np_}+{dp})'
        formula = (f"={cl}{cb['training_per_m']}+{inf_owned}+{infra_owned}+"
                   f"{cl}{cb['rd_per_m']}+{cl}{cb['annual_depr']}/{cl}{cb['lab_tokens']}")
        cell = ws.cell(row=r, column=col, value=formula)
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.number_format = '$#,##0.00'
        cell.fill = ECONOMIC_TOTAL
    cb['total_economic'] = r; r += 2

    # ---- Layer share of total ----
    ws.cell(row=r, column=1, value="Layer share of total lab cash cost (Mid)").font = BLACK_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
    for label, key in [("Training %", 'training_per_m'),
                        ("Inference %", 'inference_per_m'),
                        ("Infrastructure %", 'infra_per_m'),
                        ("R&D %", 'rd_per_m'),
                        ("Capex (lab borne) %", 'capex_per_m')]:
        ws.cell(row=r, column=1, value=label).font = BLACK
        cell = ws.cell(row=r, column=4, value=f"=D{cb[key]}/D{cb['total']}")
        cell.font = BLACK
        cell.number_format = '0.0%'
        r += 1

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 42
    for c in ['C', 'D', 'E']:
        ws.column_dimensions[c].width = 14

    return cb


# ============================================================================
# REVENUE BUILD SHEET
# ============================================================================

def build_revenue_build(wb, rows):
    """Build the Revenue Build sheet. Returns dict of key row numbers."""
    ws = wb.create_sheet("Revenue Build")

    ws['A1'] = "Revenue Build"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    write_header_row(ws, 3, ['Component', 'Formula Logic', 'Low', 'Mid', 'High'])

    rv = {}
    def aref(key, col):
        return assumption_ref(rows, key, col)

    r = 4
    write_section_label(ws, r, "Consumer subscription revenue"); r += 1

    ws.cell(row=r, column=1, value="Paid subscribers (M)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('subs', scen)}")
        cell.font = GREEN
        cell.number_format = '#,##0.0'
    rv['subs'] = r; r += 1

    ws.cell(row=r, column=1, value="Avg sub price ($/mo)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('sub_price', scen)}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
    rv['price'] = r; r += 1

    ws.cell(row=r, column=1, value="Annual consumer revenue ($M)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="Subs x price x 12").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{rv['subs']}*{cl}{rv['price']}*12")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0'
        cell.fill = TOTAL_FILL
    rv['consumer_rev'] = r; r += 2

    ws.cell(row=r, column=1, value="Consumer output tokens (annual, trillions)").font = BLACK
    ws.cell(row=r, column=2, value="Subs x queries/mo x tokens/q x 12").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{rv['subs']}*1000000*{aref('queries_pm', scen)}*"
                             f"{aref('tokens_per_q', scen)}*12/1000000000000")
        cell.font = BLACK
        cell.number_format = '#,##0.0'
    rv['consumer_tokens'] = r; r += 1

    ws.cell(row=r, column=1, value="Consumer revenue per M output tokens").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{rv['consumer_rev']}/{cl}{rv['consumer_tokens']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00'
        cell.fill = TOTAL_FILL
    rv['consumer_yield'] = r; r += 2

    write_section_label(ws, r, "API and contract revenue"); r += 1

    for label, key, target in [("API revenue ($M)", 'api_rev', 'api'),
                                ("Enterprise revenue ($M)", 'ent_rev', 'ent'),
                                ("Government revenue ($M)", 'gov_rev', 'gov')]:
        ws.cell(row=r, column=1, value=label).font = BLACK
        for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
            cell = ws.cell(row=r, column=col, value=f"={aref(key, scen)}")
            cell.font = GREEN
            cell.number_format = '$#,##0'
        rv[target] = r
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="TOTAL annual revenue ($M)").font = Font(name='Arial', size=12, bold=True)
    ws.cell(row=r, column=1).fill = PRIMARY_TOTAL
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{rv['consumer_rev']}+{cl}{rv['api']}+{cl}{rv['ent']}+{cl}{rv['gov']}")
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.number_format = '$#,##0'
        cell.fill = PRIMARY_TOTAL
    rv['total'] = r; r += 2

    ws.cell(row=r, column=1, value="Revenue mix (Mid)").font = BLACK_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
    for label, src in [("Consumer %", 'consumer_rev'), ("API %", 'api'),
                        ("Enterprise %", 'ent'), ("Government %", 'gov')]:
        ws.cell(row=r, column=1, value=label).font = BLACK
        cell = ws.cell(row=r, column=4, value=f"=D{rv[src]}/D{rv['total']}")
        cell.font = BLACK
        cell.number_format = '0.0%'
        r += 1

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 32
    for c in ['C', 'D', 'E']:
        ws.column_dimensions[c].width = 14

    return rv


# ============================================================================
# UNIT ECONOMICS SHEET
# ============================================================================

def build_unit_economics(wb, rows, cb, rv):
    """Build the Unit Economics sheet."""
    ws = wb.create_sheet("Unit Economics")

    ws['A1'] = "Unit Economics: Cost vs Revenue"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    write_header_row(ws, 3, ['Metric', 'Notes', 'Low', 'Mid', 'High'])

    ue = {}
    def aref(key, col):
        return assumption_ref(rows, key, col)

    r = 4
    ws.cell(row=r, column=1, value="Cost per M tokens (lab cash)").font = BLACK
    ws.cell(row=r, column=2, value="Capex borne by lab only").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Cost Build'!{cl}{cb['total']}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
    ue['cost'] = r; r += 1

    ws.cell(row=r, column=1, value="Cost per M tokens (full economic)").font = BLACK
    ws.cell(row=r, column=2, value="Includes hyperscaler-borne capex").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Cost Build'!{cl}{cb['total_economic']}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
    ue['cost_econ'] = r; r += 1

    ws.cell(row=r, column=1, value="Consumer revenue per M tokens").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Revenue Build'!{cl}{rv['consumer_yield']}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
    ue['cons_yield'] = r; r += 1

    ws.cell(row=r, column=1, value="Consumer margin vs lab cash cost").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR(({cl}{ue['cons_yield']}-{cl}{ue['cost']})/{cl}{ue['cons_yield']},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%;(0.0%);-'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Consumer margin vs full economic cost").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR(({cl}{ue['cons_yield']}-{cl}{ue['cost_econ']})/{cl}{ue['cons_yield']},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%;(0.0%);-'
        cell.fill = TOTAL_FILL
    r += 2

    write_section_label(ws, r, "Per-query view (consumer)"); r += 1

    ws.cell(row=r, column=1, value="Avg output tokens per query").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('tokens_per_q', scen)}")
        cell.font = GREEN
        cell.number_format = '#,##0'
    ue['tpq'] = r; r += 1

    for label, src_key, dest_key in [("Cost per query - lab cash ($)", 'cost', 'cost_q'),
                                       ("Cost per query - full economic ($)", 'cost_econ', 'cost_q_econ'),
                                       ("Revenue per query ($, consumer)", 'cons_yield', 'rev_q')]:
        ws.cell(row=r, column=1, value=label).font = BLACK
        for col in [3, 4, 5]:
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col, value=f"={cl}{ue[src_key]}*{cl}{ue['tpq']}/1000000")
            cell.font = BLACK
            cell.number_format = '$#,##0.0000'
        ue[dest_key] = r
        r += 1

    ws.cell(row=r, column=1, value="Margin per query - lab cash ($)").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{ue['rev_q']}-{cl}{ue['cost_q']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.0000;($#,##0.0000);-'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Margin per query - full economic ($)").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{ue['rev_q']}-{cl}{ue['cost_q_econ']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.0000;($#,##0.0000);-'
        cell.fill = TOTAL_FILL
    r += 2

    write_section_label(ws, r, "Company-wide view"); r += 1

    ws.cell(row=r, column=1, value="Total annual revenue ($M)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Revenue Build'!{cl}{rv['total']}")
        cell.font = GREEN
        cell.number_format = '$#,##0'
    ue['tot_rev'] = r; r += 1

    ws.cell(row=r, column=1, value="Total annual cost - lab cash ($M)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{ue['cost']}*'Cost Build'!{cl}{cb['lab_tokens']}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    ue['tot_cost'] = r; r += 1

    ws.cell(row=r, column=1, value="Total annual cost - full economic ($M)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{ue['cost_econ']}*'Cost Build'!{cl}{cb['lab_tokens']}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    ue['tot_cost_econ'] = r; r += 1

    ws.cell(row=r, column=1, value="Operating margin - lab cash ($M)").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{ue['tot_rev']}-{cl}{ue['tot_cost']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0;($#,##0);-'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Operating margin - full economic ($M)").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{ue['tot_rev']}-{cl}{ue['tot_cost_econ']}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0;($#,##0);-'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Operating margin % - full economic").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR(({cl}{ue['tot_rev']}-{cl}{ue['tot_cost_econ']})/{cl}{ue['tot_rev']},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%;(0.0%);-'
        cell.fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 32
    for c in ['C', 'D', 'E']:
        ws.column_dimensions[c].width = 14


# ============================================================================
# SENSITIVITY SHEET
# ============================================================================

def build_sensitivity(wb, cb):
    """Build the Sensitivity sheet."""
    ws = wb.create_sheet("Sensitivity")

    ws['A1'] = "Sensitivity: Cost per M Tokens (Mid base, lab cash basis)"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:G1')

    ws['A2'] = "Each row varies one driver while holding others at Mid."
    ws['A2'].font = NOTE
    ws.merge_cells('A2:G2')

    write_header_row(ws, 4, ['Driver', 'Direction', '-50%', '-25%', 'Base (Mid)', '+25%', '+50%'])

    base = f"'Cost Build'!D{cb['total']}"
    tr = f"'Cost Build'!D{cb['training_per_m']}"
    inf = f"'Cost Build'!D{cb['inference_per_m']}"
    infra = f"'Cost Build'!D{cb['infra_per_m']}"
    rd = f"'Cost Build'!D{cb['rd_per_m']}"
    cap = f"'Cost Build'!D{cb['capex_per_m']}"

    r = 5
    sensitivity_rows = [
        ("Training run cost", "Affects training layer",
         lambda m: f"{base}-{tr}+{tr}*{m}"),
        ("Lab annual tokens served", "Affects training, R&D, capex (inverse)",
         lambda m: f"{base}-{tr}-{rd}-{cap}+({tr}+{rd}+{cap})/{m}"),
        ("GPU $/hr", "Affects inference + infra",
         lambda m: f"{base}-{inf}-{infra}+({inf}+{infra})*{m}"),
        ("Tokens per GPU-hr", "Inference efficiency (inverse)",
         lambda m: f"{base}-{inf}-{infra}+({inf}+{infra})/{m}"),
        ("Annual R&D spend", "Affects R&D layer",
         lambda m: f"{base}-{rd}+{rd}*{m}"),
        ("Useful life of model", "Longer life = more amortization base",
         lambda m: f"{base}-{tr}+{tr}/{m}"),
        ("Datacenter capacity (MW)", "Affects capex layer",
         lambda m: f"{base}-{cap}+{cap}*{m}"),
        ("GPU count", "Affects capex layer",
         lambda m: f"{base}-{cap}+{cap}*{m}"),
    ]

    for label, direction, formula_fn in sensitivity_rows:
        ws.cell(row=r, column=1, value=label).font = BLACK
        ws.cell(row=r, column=2, value=direction).font = BLACK
        for col, m in zip([3, 4, 5, 6, 7], [0.5, 0.75, 1.0, 1.25, 1.5]):
            cell = ws.cell(row=r, column=col, value=f"={formula_fn(m)}")
            cell.font = BLACK
            cell.number_format = '$#,##0.00'
        r += 1

    # Special case: capex ownership %
    ws.cell(row=r, column=1, value="Capex ownership %").font = BLACK
    ws.cell(row=r, column=2, value="See Build vs Rent sheet for full toggle").font = BLACK
    for col in range(3, 8):
        ws.cell(row=r, column=col, value="See Build vs Rent sheet").font = NOTE
    r += 2

    write_section_label(ws, r, "Read: which lever moves cost most?", span=7); r += 1
    write_takeaway_block(ws, r, [
        "If a row's spread is wide, that driver dominates true cost.",
        "Lab annual tokens now affects three layers (training, R&D, capex) so it is the most leveraged driver.",
        "Capex layer makes datacenter capacity and GPU count first-order drivers, not rounding errors.",
    ], span=7)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 38
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[c].width = 13


# ============================================================================
# BUILD VS RENT SHEET
# ============================================================================

def build_build_vs_rent(wb, rows, cb):
    """Build the Build vs Rent sheet."""
    ws = wb.create_sheet("Build vs Rent")

    ws['A1'] = "Build vs Rent: Capex Allocation Toggle"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:G1')

    ws['A2'] = ("Shows total cost per M tokens at different ownership mixes. "
                "Base case uses Mid scenario assumptions.")
    ws['A2'].font = NOTE
    ws.merge_cells('A2:G2')

    ws['A4'] = "Key insight"
    ws['A4'].font = SECTION
    ws['A5'] = ("Building your own datacenter front-loads capex but lowers ongoing inference cost. "
                "Renting from hyperscalers shifts capex burden to them, but they pass it through "
                "as higher GPU $/hr. Total ECONOMIC cost should be similar across mixes — it's a "
                "question of who carries the risk and balance sheet.")
    ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A5:G7')

    write_header_row(ws, 9, ['Ownership %', '0% (all rent)', '25%', '50% (mixed)',
                              '75%', '100% (all own)'])

    def aref(key, col):
        return assumption_ref(rows, key, col)

    # Mid-scenario references
    own_rate = aref('gpu_hr_owned', 'D')
    rent_rate = aref('gpu_hr_rented', 'D')
    tph = aref('tokens_per_hr', 'D')
    io_r = aref('io_ratio', 'D')
    in_f = aref('input_factor', 'D')
    pp = aref('power_pct', 'D')
    np_ = aref('net_pct', 'D')
    dp = aref('dc_pct', 'D')
    ann_depr = f"'Cost Build'!D{cb['annual_depr']}"
    lab_tok = f"'Cost Build'!D{cb['lab_tokens']}"
    training_pm = f"'Cost Build'!D{cb['training_per_m']}"
    rd_pm = f"'Cost Build'!D{cb['rd_per_m']}"

    r = 10
    ws.cell(row=r, column=1, value="Effective GPU $/hr").font = BLACK
    for col, own_pct in zip([2, 3, 4, 5, 6], [0.0, 0.25, 0.5, 0.75, 1.0]):
        cell = ws.cell(row=r, column=col, value=f"={own_pct}*{own_rate}+(1-{own_pct})*{rent_rate}")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    gpu_hr_row = r; r += 1

    ws.cell(row=r, column=1, value="Inference $/M tokens (output+input)").font = BLACK
    for col in range(2, 7):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{gpu_hr_row}*1000000/{tph}*(1+{io_r}*{in_f})")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    inf_pm_row = r; r += 1

    ws.cell(row=r, column=1, value="Infra overhead $/M tokens").font = BLACK
    for col in range(2, 7):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{inf_pm_row}*({pp}+{np_}+{dp})")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    infra_pm_row = r; r += 1

    ws.cell(row=r, column=1, value="Capex $/M tokens (lab cash)").font = BLACK
    for col, own_pct in zip([2, 3, 4, 5, 6], [0.0, 0.25, 0.5, 0.75, 1.0]):
        cell = ws.cell(row=r, column=col, value=f"={ann_depr}*{own_pct}/{lab_tok}")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    capex_pm_row = r; r += 1

    ws.cell(row=r, column=1, value="Training + R&D $/M tokens").font = BLACK
    for col in range(2, 7):
        cell = ws.cell(row=r, column=col, value=f"={training_pm}+{rd_pm}")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    fixed_pm_row = r; r += 1

    ws.cell(row=r, column=1, value="TOTAL lab cash cost $/M tokens").font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=r, column=1).fill = PRIMARY_TOTAL
    for col in range(2, 7):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{inf_pm_row}+{cl}{infra_pm_row}+{cl}{capex_pm_row}+{cl}{fixed_pm_row}")
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.number_format = '$#,##0.00'
        cell.fill = PRIMARY_TOTAL
    total_lab_row = r; r += 1

    ws.cell(row=r, column=1, value="TOTAL economic cost $/M tokens").font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=r, column=1).fill = ECONOMIC_TOTAL
    ws.cell(row=r, column=7, value="(Should be ~constant across mixes)").font = NOTE
    # Economic cost: owned-rate inference + infra on that + full depreciation + training + R&D
    inf_owned = f'({own_rate}*1000000/{tph}*(1+{io_r}*{in_f}))'
    infra_owned = f'{inf_owned}*({pp}+{np_}+{dp})'
    for col in range(2, 7):
        cell = ws.cell(row=r, column=col,
                       value=f"={inf_owned}+{infra_owned}+{ann_depr}/{lab_tok}+{training_pm}+{rd_pm}")
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.number_format = '$#,##0.00'
        cell.fill = ECONOMIC_TOTAL
    r += 2

    write_section_label(ws, r, "Capex burden split", span=6); r += 1

    ws.cell(row=r, column=1, value="Annual depreciation TOTAL ($M)").font = BLACK
    for col in range(2, 7):
        cell = ws.cell(row=r, column=col, value=f"={ann_depr}")
        cell.font = GREEN
        cell.number_format = '$#,##0'
    r += 1

    ws.cell(row=r, column=1, value="Borne by lab ($M/yr)").font = BLACK
    for col, own_pct in zip([2, 3, 4, 5, 6], [0.0, 0.25, 0.5, 0.75, 1.0]):
        cell = ws.cell(row=r, column=col, value=f"={ann_depr}*{own_pct}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    r += 1

    ws.cell(row=r, column=1, value="Borne by hyperscaler ($M/yr)").font = BLACK
    for col, own_pct in zip([2, 3, 4, 5, 6], [0.0, 0.25, 0.5, 0.75, 1.0]):
        cell = ws.cell(row=r, column=col, value=f"={ann_depr}*(1-{own_pct})")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    r += 1

    ws.cell(row=r, column=1, value="Total capex commitment, undepreciated ($B)").font = BLACK_BOLD
    cell = ws.cell(row=r, column=2,
                   value=f"=('Cost Build'!D{cb['dc_capex']}+'Cost Build'!D{cb['gpu_capex']}+"
                         f"'Cost Build'!D{cb['power_capex']})/1000")
    cell.font = BLACK_BOLD
    cell.number_format = '$#,##0.0'
    r += 2

    write_section_label(ws, r, "What this means", span=6); r += 1
    write_takeaway_block(ws, r, [
        "At 0% ownership: Lab pays only for tokens consumed via rental rates. Hyperscaler bears full capex risk.",
        "At 100% ownership: Lab takes all capex on its own balance sheet, pays only operating costs in inference.",
        "The Total ECONOMIC cost row should be roughly constant (orange row). What changes is WHO carries the risk.",
        "Take-or-pay contracts effectively make 'rent' look like 'own' from a risk perspective even though it sits off-balance-sheet.",
        "Capex is the elephant in the room: total commitment is far larger than any single year's revenue.",
    ], span=6, row_height=28)

    ws.column_dimensions['A'].width = 38
    for c in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[c].width = 16


# ============================================================================
# COMPETITOR STACKS — PRE ANNOUNCEMENT
# ============================================================================

def build_stacks_pre(wb, cb, rv):
    """Build the pre-announcement competitor comparison sheet."""
    ws = wb.create_sheet("Stacks - Pre Announcement")

    # Banner explaining this is the legacy view
    ws['A1'] = ("PRE-ANNOUNCEMENT (early 2026): Anthropic modeled as Amazon-stack proxy. "
                "See Stacks - Post Announcement for updated framing after the April/May 2026 "
                "Anthropic compute deals (Google $40B + TPUs, xAI Colossus 1, expanded Amazon).")
    ws['A1'].font = WARNING
    ws['A1'].fill = WARNING_FILL
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 36
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='center')

    ws['A2'] = "Three-Way Competitive Cost Comparison"
    ws['A2'].font = TITLE
    ws.merge_cells('A2:E2')

    ws['A3'] = ("Multipliers applied to Mid scenario base cost. 1.0x = baseline lab cost. "
                "<1.0x = cost advantage. >1.0x = cost disadvantage.")
    ws['A3'].font = NOTE
    ws.merge_cells('A3:E3')

    r = 5
    write_section_label(ws, r, "The three vertical stacks"); r += 1

    stacks = [
        ("Google", "Search/Ads -> TPUs -> Datacenters -> Gemini -> Workspace/Android/Search distribution", GOOGLE_FILL),
        ("Microsoft + OpenAI", "Azure -> Nvidia GPUs -> Datacenters -> GPT -> Office/Windows/GitHub distribution", OAI_FILL),
        ("Amazon + Anthropic", "AWS -> Trainium/Nvidia -> Datacenters -> Claude -> AWS Bedrock distribution", ANTH_FILL),
    ]
    for name, desc, fill in stacks:
        ws.cell(row=r, column=1, value=name).font = BLACK_BOLD
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=2, value=desc).font = BLACK
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    write_section_label(ws, r, "Cost layer multipliers vs baseline lab (editable - blue cells)"); r += 1

    write_header_row(ws, r, ['Cost Layer', 'Google', 'MSFT+OpenAI', 'Amazon+Anthropic', 'Notes'])
    r += 1

    mult_rows_start = r
    multipliers = [
        ("Training cost", 1.0, 1.0, 1.0, "Roughly comparable across labs at the frontier"),
        ("Inference compute (silicon)", 0.55, 1.0, 0.85, "Google: TPU vs Nvidia 70% margin. Anthropic: some Trainium offset. OpenAI: pays full Nvidia tax."),
        ("Infrastructure overhead", 0.65, 1.0, 0.95, "Google has 25 yrs of hyperscale ops experience. MSFT/AWS catching up but trail."),
        ("R&D / people / data", 0.85, 1.0, 1.0, "Google has captive data (Search/YouTube). All have similar talent costs."),
        ("Capex (effective burden)", 0.7, 1.0, 1.0, "Google self-funds from $90B+ cash flow. MSFT/Amazon also strong but capex is dedicated AI buildout."),
        ("Distribution / CAC", 0.3, 0.6, 0.6, "Google: zero CAC via Search/Workspace. MSFT/Amazon: leverage existing enterprise relationships."),
    ]
    for label, g, oai, anth, note in multipliers:
        ws.cell(row=r, column=1, value=label).font = BLACK
        for col, val, fill in zip([2, 3, 4], [g, oai, anth], [GOOGLE_FILL, OAI_FILL, ANTH_FILL]):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = BLUE
            cell.number_format = '0.00"x"'
            cell.alignment = CENTER
            cell.fill = fill
        ws.cell(row=r, column=5, value=note).font = NOTE
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True)
        r += 1

    ws.cell(row=r, column=1, value="(Distribution shown for reference - it affects revenue side, not the cost layers we sum below)").font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 2

    write_section_label(ws, r, "Resulting cost per M tokens (Mid scenario, 50% capex ownership)"); r += 1

    write_header_row(ws, r, ['Cost Layer', 'Baseline lab', 'Google', 'MSFT+OpenAI', 'Amazon+Anthropic'])
    r += 1

    # Cost rows referencing Cost Build with multipliers from this sheet
    layer_specs = [
        ("Training", f"'Cost Build'!D{cb['training_per_m']}", mult_rows_start),
        ("Inference compute", f"'Cost Build'!D{cb['inference_per_m']}", mult_rows_start + 1),
        ("Infrastructure", f"'Cost Build'!D{cb['infra_per_m']}", mult_rows_start + 2),
        ("R&D / people / data", f"'Cost Build'!D{cb['rd_per_m']}", mult_rows_start + 3),
        ("Capex (lab cash)", f"'Cost Build'!D{cb['capex_per_m']}", mult_rows_start + 4),
    ]
    cost_rows_start = r
    for label, source, mult_row in layer_specs:
        ws.cell(row=r, column=1, value=label).font = BLACK
        # Baseline (B): just the source value
        cell = ws.cell(row=r, column=2, value=f"={source}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
        # C=Google (uses B mult col), D=OAI (uses C mult col), E=Anth (uses D mult col)
        for col, mult_col in zip([3, 4, 5], ['B', 'C', 'D']):
            cell = ws.cell(row=r, column=col, value=f"={source}*{mult_col}{mult_row}")
            cell.font = BLACK
            cell.number_format = '$#,##0.00'
        r += 1

    # Total
    ws.cell(row=r, column=1, value="TOTAL cost per M tokens (lab cash)").font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=r, column=1).fill = PRIMARY_TOTAL
    for col in range(2, 6):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"=SUM({cl}{cost_rows_start}:{cl}{r-1})")
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.number_format = '$#,##0.00'
        cell.fill = PRIMARY_TOTAL
    total_cost_row = r; r += 1

    # vs baseline %
    ws.cell(row=r, column=1, value="Cost vs baseline (lower = advantage)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="100%").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = '0%'
    ws.cell(row=r, column=2).alignment = CENTER
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{total_cost_row}/B{total_cost_row}")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%'
        cell.alignment = CENTER
        cell.fill = TOTAL_FILL
    r += 2

    # Margin scenarios
    write_section_label(ws, r, "What this means in practice"); r += 1
    write_header_row(ws, r, ['Scenario', 'Google margin', 'OpenAI margin', 'Anthropic margin', 'Implication'])
    r += 1

    consumer_yield_ref = f"'Revenue Build'!D{rv['consumer_yield']}"
    scenarios = [
        ("Consumer subs (~$133/M revenue yield)",
         f"=({consumer_yield_ref}-C{total_cost_row})/{consumer_yield_ref}",
         f"=({consumer_yield_ref}-D{total_cost_row})/{consumer_yield_ref}",
         f"=({consumer_yield_ref}-E{total_cost_row})/{consumer_yield_ref}",
         "Google can price-cut consumer aggressively without losing money"),
        ("API at $30/M output (mid market)",
         f"=(30-C{total_cost_row})/30",
         f"=(30-D{total_cost_row})/30",
         f"=(30-E{total_cost_row})/30",
         "At commodity pricing, OpenAI/Anthropic likely lose money. Google barely positive."),
        ("API at $75/M output (frontier premium)",
         f"=(75-C{total_cost_row})/75",
         f"=(75-D{total_cost_row})/75",
         f"=(75-E{total_cost_row})/75",
         "Premium pricing is where standalone labs survive - if they have differentiation"),
    ]
    for name, g, oai, anth, impl in scenarios:
        ws.cell(row=r, column=1, value=name).font = BLACK
        for col, formula in zip([2, 3, 4], [g, oai, anth]):
            cell = ws.cell(row=r, column=col, value=formula)
            cell.font = BLACK
            cell.number_format = '0.0%;(0.0%);-'
            cell.alignment = CENTER
        ws.cell(row=r, column=5, value=impl).font = NOTE
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    write_section_label(ws, r, "Strategic takeaways"); r += 1
    write_takeaway_block(ws, r, [
        "Google's structural cost advantage is real and probably permanent. Roughly 35-50% lower all-in cost per token.",
        "MSFT+OpenAI and Amazon+Anthropic are competitive only by absorbing the cost gap into hyperscaler P&Ls. They cannot win as standalones.",
        "At commodity API prices, only Google has clear positive margin. Standalone labs need premium positioning or scale subsidies.",
        "Google's weakness: organizational and product execution. They've fumbled product launches before. The bull case for OpenAI/Anthropic is that capability differentiation outruns Google's cost advantage.",
        "The three-way structure favors Google in steady state. The standalone labs need to win NOW - before commoditization kicks in - or accept structural subordination.",
        "This explains why OpenAI is pushing so hard into consumer (ChatGPT, Atlas browser) and Anthropic into enterprise/code (Claude Code, Bedrock). Both are trying to escape pure API commodity competition.",
    ])

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 38

    return {'total_cost_row': total_cost_row}


# ============================================================================
# COMPETITOR STACKS — POST ANNOUNCEMENT
# ============================================================================

def build_stacks_post(wb, cb, rv, pre):
    """Build the post-announcement competitor comparison sheet."""
    ws = wb.create_sheet("Stacks - Post Announcement")

    ws['A1'] = "POST-ANNOUNCEMENT: Multi-Stack Competitive Reality (May 2026)"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    ws['A2'] = ("Reflects: Anthropic-Google $40B+TPU deal (Apr 2026), Anthropic-xAI Colossus 1 deal "
                "(May 2026), Amazon expansion to ~$100B / 5GW Trainium. Anthropic now multi-stack "
                "across Trainium, TPU, and Nvidia.")
    ws['A2'].font = NOTE
    ws['A2'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A2:E3')
    ws.row_dimensions[2].height = 30

    r = 5
    write_section_label(ws, r, "Revised competitive structure"); r += 1

    stacks = [
        ("Google", "Vertically integrated AND major arms dealer. Sells TPU capacity to Anthropic at scale. "
                  "Trades some moat for cloud revenue and TPU validation as Nvidia alternative.", GOOGLE_FILL),
        ("OpenAI (single-stack)", "Most strategically constrained: locked into Microsoft/Azure/Nvidia. "
                                   "Cannot capture TPU economics. Cannot dual-source. Largest revenue but narrowest supplier base.", OAI_FILL),
        ("Anthropic (multi-stack)", "Runs on Trainium (AWS), TPU (Google), Nvidia (xAI/CoreWeave). "
                                     "Captures portion of TPU cost advantage via Google deal. "
                                     "Portfolio diversification as strategic moat.", ANTH_FILL),
    ]
    for name, desc, fill in stacks:
        ws.cell(row=r, column=1, value=name).font = BLACK_BOLD
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=2, value=desc).font = BLACK
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 42
        r += 1

    r += 1
    write_section_label(ws, r, "Updated cost layer multipliers vs baseline lab (post-announcement)"); r += 1
    write_header_row(ws, r, ['Cost Layer', 'Google', 'OpenAI (MSFT)', 'Anthropic (multi)',
                              'Change vs pre-announcement'])
    r += 1

    mult_rows_start = r
    multipliers = [
        ("Training cost", 1.0, 1.0, 1.0,
         "Unchanged - frontier training comparable across labs"),
        ("Inference compute (silicon)", 0.55, 1.0, 0.72,
         "Anthropic 0.85 -> 0.72: TPU access flows through partial cost advantage"),
        ("Infrastructure overhead", 0.65, 1.0, 0.85,
         "Anthropic 0.95 -> 0.85: TPU + Trainium + scale = better infra leverage"),
        ("R&D / people / data", 0.85, 1.0, 1.0, "Unchanged"),
        ("Capex (effective burden)", 0.7, 1.0, 0.85,
         "Anthropic 1.0 -> 0.85: multi-vendor sourcing reduces single-supplier markup"),
        ("Distribution / CAC", 0.3, 0.55, 0.5,
         "OpenAI 0.6 -> 0.55 (consumer scale). Anthropic 0.6 -> 0.5 (Bedrock+Vertex+Azure)"),
    ]
    for label, g, oai, anth, note in multipliers:
        ws.cell(row=r, column=1, value=label).font = BLACK
        for col, val, fill in zip([2, 3, 4], [g, oai, anth], [GOOGLE_FILL, OAI_FILL, ANTH_FILL]):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = BLUE
            cell.number_format = '0.00"x"'
            cell.alignment = CENTER
            cell.fill = fill
        ws.cell(row=r, column=5, value=note).font = NOTE
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 28
        r += 1

    ws.cell(row=r, column=1,
            value="(Distribution shown for reference - affects revenue side, not summed in cost total)").font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 2

    write_section_label(ws, r, "Resulting cost per M tokens (Mid scenario, 50% capex ownership)"); r += 1
    write_header_row(ws, r, ['Cost Layer', 'Baseline lab', 'Google', 'OpenAI', 'Anthropic'])
    r += 1

    layer_specs = [
        ("Training", f"'Cost Build'!D{cb['training_per_m']}", mult_rows_start),
        ("Inference compute", f"'Cost Build'!D{cb['inference_per_m']}", mult_rows_start + 1),
        ("Infrastructure", f"'Cost Build'!D{cb['infra_per_m']}", mult_rows_start + 2),
        ("R&D / people / data", f"'Cost Build'!D{cb['rd_per_m']}", mult_rows_start + 3),
        ("Capex (lab cash)", f"'Cost Build'!D{cb['capex_per_m']}", mult_rows_start + 4),
    ]
    cost_rows_start = r
    for label, source, mult_row in layer_specs:
        ws.cell(row=r, column=1, value=label).font = BLACK
        cell = ws.cell(row=r, column=2, value=f"={source}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
        for col, mult_col in zip([3, 4, 5], ['B', 'C', 'D']):
            cell = ws.cell(row=r, column=col, value=f"={source}*{mult_col}{mult_row}")
            cell.font = BLACK
            cell.number_format = '$#,##0.00'
        r += 1

    ws.cell(row=r, column=1, value="TOTAL cost per M tokens (lab cash)").font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=r, column=1).fill = PRIMARY_TOTAL
    for col in range(2, 6):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"=SUM({cl}{cost_rows_start}:{cl}{r-1})")
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.number_format = '$#,##0.00'
        cell.fill = PRIMARY_TOTAL
    total_cost_row = r; r += 1

    ws.cell(row=r, column=1, value="Cost vs baseline (lower = advantage)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="100%").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = '0%'
    ws.cell(row=r, column=2).alignment = CENTER
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{total_cost_row}/B{total_cost_row}")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%'
        cell.alignment = CENTER
        cell.fill = TOTAL_FILL
    r += 2

    # Side-by-side comparison vs pre-announcement
    write_section_label(ws, r, "Side-by-side: pre vs post announcement"); r += 1
    write_header_row(ws, r, ['Metric', 'Baseline lab', 'Google', 'OpenAI', 'Anthropic'])
    r += 1

    pre_total_row = pre['total_cost_row']

    ws.cell(row=r, column=1, value="Pre-announcement total $/M").font = BLACK
    for col, pre_col in zip([2, 3, 4, 5], ['B', 'C', 'D', 'E']):
        cell = ws.cell(row=r, column=col,
                       value=f"='Stacks - Pre Announcement'!{pre_col}{pre_total_row}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
    pre_total_disp_row = r; r += 1

    ws.cell(row=r, column=1, value="Post-announcement total $/M").font = BLACK
    for col in [2, 3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{total_cost_row}")
        cell.font = BLACK
        cell.number_format = '$#,##0.00'
    post_total_disp_row = r; r += 1

    ws.cell(row=r, column=1, value="Change ($)").font = BLACK_BOLD
    for col in [2, 3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{post_total_disp_row}-{cl}{pre_total_disp_row}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(row=r, column=5).fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Change (%)").font = BLACK_BOLD
    for col in [2, 3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR(({cl}{post_total_disp_row}-{cl}{pre_total_disp_row})/{cl}{pre_total_disp_row},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%;(0.0%);-'
    ws.cell(row=r, column=5).fill = TOTAL_FILL
    r += 2

    # Margin scenarios
    write_section_label(ws, r, "Margin at different price points (post-announcement)"); r += 1
    write_header_row(ws, r, ['Scenario', 'Google margin', 'OpenAI margin', 'Anthropic margin', 'Implication'])
    r += 1

    consumer_yield_ref = f"'Revenue Build'!D{rv['consumer_yield']}"
    scenarios = [
        ("Consumer subs (~$133/M revenue yield)",
         f"=({consumer_yield_ref}-C{total_cost_row})/{consumer_yield_ref}",
         f"=({consumer_yield_ref}-D{total_cost_row})/{consumer_yield_ref}",
         f"=({consumer_yield_ref}-E{total_cost_row})/{consumer_yield_ref}",
         "Anthropic now closer to Google margin via TPU access"),
        ("API at $30/M output (mid market)",
         f"=(30-C{total_cost_row})/30",
         f"=(30-D{total_cost_row})/30",
         f"=(30-E{total_cost_row})/30",
         "Still loss-making for all at mid pricing - capex burden dominates"),
        ("API at $75/M output (frontier premium)",
         f"=(75-C{total_cost_row})/75",
         f"=(75-D{total_cost_row})/75",
         f"=(75-E{total_cost_row})/75",
         "Anthropic gap to Google narrowed but OpenAI now most disadvantaged"),
    ]
    for name, g, oai, anth, impl in scenarios:
        ws.cell(row=r, column=1, value=name).font = BLACK
        for col, formula in zip([2, 3, 4], [g, oai, anth]):
            cell = ws.cell(row=r, column=col, value=formula)
            cell.font = BLACK
            cell.number_format = '0.0%;(0.0%);-'
            cell.alignment = CENTER
        ws.cell(row=r, column=5, value=impl).font = NOTE
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    write_section_label(ws, r, "Revised strategic takeaways"); r += 1
    write_takeaway_block(ws, r, [
        "Anthropic's cost gap vs Google narrowed substantially. From ~99% of baseline (pre) to ~88% (post). Still disadvantaged vs Google but materially closer.",
        "OpenAI now the most strategically constrained of the three: locked into Nvidia + Microsoft, cannot capture TPU economics, single-vendor risk on silicon.",
        "Google's vertical integration moat is partially monetized - they sell TPU capacity to a competitor for cloud revenue. Smart short-term, dilutive long-term.",
        "Multi-stack diversification has emerged as a real competitive strategy. Anthropic running on Trainium + TPU + Nvidia is a new pattern.",
        "Demand growth (Anthropic $9B -> $30B run-rate in 4 months) suggests revenue may be compounding fast enough to justify capex commitments. Bear case weakened.",
        "xAI/SpaceX renting out Colossus to Anthropic is a tell that even purpose-built capacity is outrunning native demand - some 'overbuild' risk is real.",
        "Counterparty risk language in the Broadcom SEC filing ('dependent on Anthropic's continued commercial success') signals suppliers see real risk in these commitments.",
    ])

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 38


# ============================================================================
# BREAKEVEN SHEET
# ============================================================================

def build_breakeven(wb, rows, cb, rv):
    """Reframe the model around what demand picture the spending implies."""
    ws = wb.create_sheet("Breakeven")

    ws['A1'] = "Breakeven at Current Cost Structure"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:G1')

    ws['A2'] = ("Inverts the model: at the current full-economic cost per million tokens, "
                "what revenue mix and scale would clear the cost stack?")
    ws['A2'].font = NOTE
    ws.merge_cells('A2:G2')

    write_header_row(ws, 4, ['Metric', 'Notes', 'Low', 'Mid', 'High'])

    r = 5
    write_section_label(ws, r, "A. Breakeven price per million tokens"); r += 1

    ws.cell(row=r, column=1, value="Full-economic cost per M tokens").font = BLACK
    ws.cell(row=r, column=2, value="Margin-neutral price target").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Cost Build'!{cl}{cb['total_economic']}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
    be_cost = r; r += 1

    ws.cell(row=r, column=1, value="Current blended consumer yield per M tokens").font = BLACK
    ws.cell(row=r, column=2, value="From Revenue Build").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Revenue Build'!{cl}{rv['consumer_yield']}")
        cell.font = GREEN
        cell.number_format = '$#,##0.00'
    be_yield = r; r += 1

    ws.cell(row=r, column=1, value="Implicit subsidy per M tokens").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="Cost minus consumer yield").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"={cl}{be_cost}-{cl}{be_yield}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Required price multiple over current yield").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"=IFERROR({cl}{be_cost}/{cl}{be_yield},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0"x"'
        cell.fill = TOTAL_FILL
    r += 2

    write_section_label(ws, r, "B. Annual token volume needed at current revenue mix"); r += 1

    ws.cell(row=r, column=1, value="Current total revenue ($M)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Revenue Build'!{cl}{rv['total']}")
        cell.font = GREEN
        cell.number_format = '$#,##0'
    be_rev = r; r += 1

    ws.cell(row=r, column=1, value="Lab tokens served today (trillions/yr)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Cost Build'!{cl}{cb['lab_tokens']}")
        cell.font = GREEN
        cell.number_format = '#,##0.0'
    be_today = r; r += 1

    ws.cell(row=r, column=1, value="Tokens needed to clear cost at today's revenue (T/yr)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="Rev$ / cost per M = M-token-units; div 1M for trillions").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{be_rev}/{cl}{be_cost},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '#,##0.0'
        cell.fill = TOTAL_FILL
    be_needed = r; r += 1

    ws.cell(row=r, column=1, value="Revenue coverage of today's volume").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="needed / served today (>100% = overfunded, <100% = revenue gap)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{be_needed}/{cl}{be_today},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Revenue gap on today's volume (T tokens/yr unfunded)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="served today - tokens revenue can clear").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{be_today}-{cl}{be_needed}")
        cell.font = BLACK_BOLD
        cell.number_format = '#,##0.0'
        cell.fill = TOTAL_FILL
    r += 2

    # ---- C. Tradeoff grid: price/M vs revenue target ----
    write_section_label(ws, r, "C. Trade-off grid: revenue per M tokens vs annual revenue ($B), Mid cost basis", span=7); r += 1
    ws.cell(row=r, column=1, value="Tokens needed (trillions/yr) to clear full-economic cost").font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); r += 1

    price_points = [20, 50, 100, 150, 253, 400]
    revenue_targets_b = [5, 20, 50, 100, 250]

    write_header_row(ws, r, ['$/M tokens'] + [f"${t}B/yr" for t in revenue_targets_b])
    r += 1

    for price in price_points:
        ws.cell(row=r, column=1, value=f"${price}").font = BLACK_BOLD
        ws.cell(row=r, column=1).alignment = RIGHT
        for col_idx, target_b in enumerate(revenue_targets_b, start=2):
            target_m = target_b * 1000
            cell = ws.cell(row=r, column=col_idx, value=f"={target_m}/{price}")
            cell.font = BLACK
            cell.number_format = '#,##0.0'
            cell.alignment = RIGHT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=f"Reading: at $50/M and $20B annual revenue target, the lab must serve {20000/50/1000000*1000000:.0f}T tokens/yr.").font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 42
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[c].width = 14


# ============================================================================
# NO-SUBSIDY PRICING SHEET
# ============================================================================

def build_no_subsidy(wb, rows, cb):
    """What consumer subscriptions would cost without subsidy."""
    ws = wb.create_sheet("No-Subsidy Pricing")

    ws['A1'] = "No-Subsidy Consumer Subscription Pricing"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    ws['A2'] = ("What a subscription would cost if priced at full-economic cost. "
                "Median tier ties to Assumptions queries_pm and tokens_per_q.")
    ws['A2'].font = NOTE
    ws.merge_cells('A2:E2')

    write_header_row(ws, 4, ['Usage tier', 'Driver', 'Low', 'Mid', 'High'])

    def aref(key, col):
        return assumption_ref(rows, key, col)

    tiers = [
        ('Light', 'light_q', 'light_tpq'),
        ('Median', 'queries_pm', 'tokens_per_q'),
        ('Power', 'power_q', 'power_tpq'),
    ]

    r = 5
    section_starts = {}
    for tier_name, q_key, tpq_key in tiers:
        write_section_label(ws, r, f"{tier_name} tier"); r += 1
        section_starts[tier_name] = r

        ws.cell(row=r, column=1, value="Queries per month").font = BLACK
        for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
            cell = ws.cell(row=r, column=col, value=f"={aref(q_key, scen)}")
            cell.font = GREEN
            cell.number_format = '#,##0'
        q_row = r; r += 1

        ws.cell(row=r, column=1, value="Output tokens per query").font = BLACK
        for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
            cell = ws.cell(row=r, column=col, value=f"={aref(tpq_key, scen)}")
            cell.font = GREEN
            cell.number_format = '#,##0'
        tpq_row = r; r += 1

        ws.cell(row=r, column=1, value="Monthly output tokens").font = BLACK
        ws.cell(row=r, column=2, value="queries x tokens/q").font = BLACK
        for col in [3, 4, 5]:
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col, value=f"={cl}{q_row}*{cl}{tpq_row}")
            cell.font = BLACK
            cell.number_format = '#,##0'
        tok_row = r; r += 1

        ws.cell(row=r, column=1, value="Monthly cost at full-economic ($/mo)").font = BLACK_BOLD
        ws.cell(row=r, column=2, value="tokens x cost per M / 1M").font = BLACK
        for col in [3, 4, 5]:
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col,
                           value=f"={cl}{tok_row}*'Cost Build'!{cl}{cb['total_economic']}/1000000")
            cell.font = BLACK_BOLD
            cell.number_format = '$#,##0.00'
            cell.fill = TOTAL_FILL
        cost_row = r; r += 1

        ws.cell(row=r, column=1, value="Current flat sub price ($/mo)").font = BLACK
        for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
            cell = ws.cell(row=r, column=col, value=f"={aref('sub_price', scen)}")
            cell.font = GREEN
            cell.number_format = '$#,##0.00'
        flat_row = r; r += 1

        ws.cell(row=r, column=1, value="Implicit monthly subsidy ($/mo)").font = BLACK_BOLD
        ws.cell(row=r, column=2, value="cost minus flat price").font = BLACK
        for col in [3, 4, 5]:
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col, value=f"={cl}{cost_row}-{cl}{flat_row}")
            cell.font = BLACK_BOLD
            cell.number_format = '$#,##0.00;($#,##0.00);-'
            cell.fill = TOTAL_FILL
        r += 1

        ws.cell(row=r, column=1, value="Effective $/M revenue at flat rate").font = BLACK
        ws.cell(row=r, column=2, value="flat price x 1M / monthly tokens").font = BLACK
        for col in [3, 4, 5]:
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col,
                           value=f"=IFERROR({cl}{flat_row}*1000000/{cl}{tok_row},0)")
            cell.font = BLACK
            cell.number_format = '$#,##0.00'
        r += 2

    r = write_takeaway_block(ws, r, [
        "Light users (50 q/mo) effectively subsidize power users when everyone pays the same flat rate.",
        "Median Mid cost-to-serve ($38/mo) is ~2x the current $20 flat rate.",
        "Power users (2k q/mo, 1k-token outputs) cost ~25x the flat rate.",
        "Implicit subsidy disappears only when (a) per-token cost falls 5-10x or (b) prices move to metered.",
    ], row_height=28)

    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 34
    for c in ['C', 'D', 'E']:
        ws.column_dimensions[c].width = 14


# ============================================================================
# IMPLICIT DEMAND SHEET
# ============================================================================

def build_implicit_demand(wb, rows, cb):
    """What demand the committed compute spend implies, in tokens/year."""
    ws = wb.create_sheet("Implicit Demand")

    ws['A1'] = "Implicit Demand from Committed Compute Spend"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    ws['A2'] = ("Inverts committed multi-year compute spend into the annual token volume it "
                "implies at current full-economic cost per million tokens.")
    ws['A2'].font = NOTE
    ws.merge_cells('A2:E2')

    write_header_row(ws, 4, ['Lab', 'Calc', 'Low', 'Mid', 'High'])

    def aref(key, col):
        return assumption_ref(rows, key, col)

    labs = [
        ('Anthropic', 'anthropic_5y_b'),
        ('OpenAI', 'openai_5y_b'),
        ('Baseline lab (xAI/Google/Meta avg)', 'baseline_lab_5y_b'),
    ]

    r = 5
    lab_token_rows = []
    for lab_name, commit_key in labs:
        write_section_label(ws, r, lab_name); r += 1

        ws.cell(row=r, column=1, value="5-year commitment ($B)").font = BLACK
        for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
            cell = ws.cell(row=r, column=col, value=f"={aref(commit_key, scen)}")
            cell.font = GREEN
            cell.number_format = '$#,##0'
        commit_row = r; r += 1

        ws.cell(row=r, column=1, value="Annual run rate ($B/yr)").font = BLACK
        ws.cell(row=r, column=2, value="commitment / years").font = BLACK
        for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col,
                           value=f"={cl}{commit_row}/{aref('commitment_years', scen)}")
            cell.font = BLACK
            cell.number_format = '$#,##0.0'
        annual_row = r; r += 1

        ws.cell(row=r, column=1, value="Full-economic cost per M tokens ($)").font = BLACK
        for col in [3, 4, 5]:
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col, value=f"='Cost Build'!{cl}{cb['total_economic']}")
            cell.font = GREEN
            cell.number_format = '$#,##0.00'
        cost_row = r; r += 1

        ws.cell(row=r, column=1, value="Implied tokens served (T/yr)").font = BLACK_BOLD
        ws.cell(row=r, column=2, value="annual $B x 1000 / cost per M / 1M").font = BLACK
        for col in [3, 4, 5]:
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col,
                           value=f"=IFERROR({cl}{annual_row}*1000/{cl}{cost_row},0)")
            cell.font = BLACK_BOLD
            cell.number_format = '#,##0.0'
            cell.fill = TOTAL_FILL
        lab_token_rows.append((lab_name, r, annual_row, commit_row))
        r += 2

    write_section_label(ws, r, "Industry total"); r += 1

    ws.cell(row=r, column=1, value="Total 5-year commitment ($B)").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        terms = "+".join(f"{cl}{cr}" for _, _, _, cr in lab_token_rows)
        cell = ws.cell(row=r, column=col, value=f"={terms}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0'
        cell.fill = TOTAL_FILL
    total_commit_row = r; r += 1

    ws.cell(row=r, column=1, value="Total annual run rate ($B/yr)").font = BLACK_BOLD
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        terms = "+".join(f"{cl}{ar}" for _, _, ar, _ in lab_token_rows)
        cell = ws.cell(row=r, column=col, value=f"={terms}")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.0'
        cell.fill = TOTAL_FILL
    total_annual_row = r; r += 1

    ws.cell(row=r, column=1, value="Total implied tokens (T/yr)").font = Font(name='Arial', size=12, bold=True)
    ws.cell(row=r, column=1).fill = PRIMARY_TOTAL
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        terms = "+".join(f"{cl}{lr}" for _, lr, _, _ in lab_token_rows)
        cell = ws.cell(row=r, column=col, value=f"={terms}")
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.number_format = '#,##0.0'
        cell.fill = PRIMARY_TOTAL
    total_tokens_row = r; r += 2

    ws.cell(row=r, column=1, value="Industry tokens served today (T/yr)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('industry_tokens', scen)}")
        cell.font = GREEN
        cell.number_format = '#,##0'
    today_row = r; r += 1

    ws.cell(row=r, column=1, value="Industry growth multiple implied").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="implied / today").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{total_tokens_row}/{cl}{today_row},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0"x"'
        cell.fill = TOTAL_FILL
    growth_row = r; r += 2

    ws.cell(row=r, column=1, value="Outputs exported to Plausibility and Capex Perspective:").font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 34
    for c in ['C', 'D', 'E']:
        ws.column_dimensions[c].width = 14

    return {
        'total_commit': total_commit_row,
        'total_annual': total_annual_row,
        'total_tokens': total_tokens_row,
        'today_tokens': today_row,
        'growth_multiple': growth_row,
    }


# ============================================================================
# PLAUSIBILITY SHEET
# ============================================================================

def build_plausibility(wb, rows, idem):
    """Three lenses on whether the implied demand is plausible."""
    ws = wb.create_sheet("Plausibility")

    ws['A1'] = "Plausibility Check: Can Implied Demand Be Real?"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    ws['A2'] = ("Three independent lenses on whether the token volumes implied by committed "
                "compute spend can plausibly be absorbed.")
    ws['A2'].font = NOTE
    ws.merge_cells('A2:E2')

    write_header_row(ws, 4, ['Metric', 'Calc', 'Low', 'Mid', 'High'])

    def aref(key, col):
        return assumption_ref(rows, key, col)

    r = 5
    write_section_label(ws, r, "Lens A: Heavy-user equivalents needed"); r += 1

    ws.cell(row=r, column=1, value="Industry implied tokens (T/yr)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Implicit Demand'!{cl}{idem['total_tokens']}")
        cell.font = GREEN
        cell.number_format = '#,##0.0'
    impl_row = r; r += 1

    ws.cell(row=r, column=1, value="Heavy user annual tokens (M/yr)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('heavy_user_tokens_yr', scen)}")
        cell.font = GREEN
        cell.number_format = '#,##0.0'
    heavy_row = r; r += 1

    ws.cell(row=r, column=1, value="Heavy-user equivalents required (millions)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="T tokens / heavy user M tokens (T/M = millions)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{impl_row}/{cl}{heavy_row},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '#,##0.0'
        cell.fill = TOTAL_FILL
    r += 2

    write_section_label(ws, r, "Lens B: Enterprise seats implied"); r += 1

    ws.cell(row=r, column=1, value="Industry annual run rate ($B/yr)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Implicit Demand'!{cl}{idem['total_annual']}")
        cell.font = GREEN
        cell.number_format = '$#,##0.0'
    annual_row = r; r += 1

    ws.cell(row=r, column=1, value="Enterprise seat price ($/seat/yr)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('seat_price', scen)}")
        cell.font = GREEN
        cell.number_format = '$#,##0'
    seat_row = r; r += 1

    ws.cell(row=r, column=1, value="Seats needed to fund commitments (millions)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="annual $B x 1B / seat price / 1M").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{annual_row}*1000000000/{cl}{seat_row}/1000000,0)")
        cell.font = BLACK_BOLD
        cell.number_format = '#,##0.0'
        cell.fill = TOTAL_FILL
    r += 2

    write_section_label(ws, r, "Lens C: Growth multiple vs today's industry"); r += 1

    ws.cell(row=r, column=1, value="Industry growth multiple implied").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="from Implicit Demand").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"='Implicit Demand'!{cl}{idem['growth_multiple']}")
        cell.font = GREEN
        cell.number_format = '0.0"x"'
        cell.fill = TOTAL_FILL
    r += 2

    r = write_takeaway_block(ws, r, [
        "Heavy-user equivalents in tens of millions imply ChatGPT-style mass adoption at higher per-user intensity than today.",
        "Enterprise seats in tens of millions implies coverage well beyond today's enterprise SaaS footprint for any single category.",
        "Growth multiples of 2-3x against a base that already grew 10x in the prior 18 months are what these commitments price in.",
    ], row_height=28)

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 36
    for c in ['C', 'D', 'E']:
        ws.column_dimensions[c].width = 14


# ============================================================================
# CAPEX PERSPECTIVE SHEET
# ============================================================================

def build_capex_perspective(wb, rows, cb, idem):
    """Frames the capex commitments against macro benchmarks."""
    ws = wb.create_sheet("Capex Perspective")

    ws['A1'] = "Capex in Perspective"
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')

    ws['A2'] = ("Two views: a single lab's footprint, and the industry-wide footprint "
                "implied by committed multi-year spend.")
    ws['A2'].font = NOTE
    ws.merge_cells('A2:E2')

    write_header_row(ws, 4, ['Metric', 'Calc', 'Low', 'Mid', 'High'])

    def aref(key, col):
        return assumption_ref(rows, key, col)

    r = 5
    write_section_label(ws, r, "View A: Per-lab footprint (one frontier lab)"); r += 1

    ws.cell(row=r, column=1, value="One-time capex total ($M)").font = BLACK
    ws.cell(row=r, column=2, value="DC + GPU + power infra").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"='Cost Build'!{cl}{cb['dc_capex']}+'Cost Build'!{cl}{cb['gpu_capex']}+'Cost Build'!{cl}{cb['power_capex']}")
        cell.font = BLACK
        cell.number_format = '$#,##0'
    capex_total = r; r += 1

    ws.cell(row=r, column=1, value="Annual depreciation ($M)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Cost Build'!{cl}{cb['annual_depr']}")
        cell.font = GREEN
        cell.number_format = '$#,##0'
    annual_depr = r; r += 1

    ws.cell(row=r, column=1, value="Datacenter capacity (MW)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('dc_capacity_mw', scen)}")
        cell.font = GREEN
        cell.number_format = '#,##0'
    dc_mw = r; r += 2

    ws.cell(row=r, column=1, value="Capex total as bps of US GDP").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="capex $M / (GDP $T x 1M) x 10000").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{capex_total}/({aref('us_gdp_t', scen)}*1000000)*10000,0)")
        cell.font = BLACK_BOLD
        cell.number_format = '#,##0.0" bps"'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Annual depreciation as % of US business capex").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="depr $M / (business capex $T x 1M)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{annual_depr}/({aref('us_business_capex_t', scen)}*1000000),0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.00%'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="DC capacity as % of US grid peak demand").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="DC MW / (grid GW x 1000)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{dc_mw}/({aref('us_grid_peak_gw', scen)}*1000),0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.00%'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Annual depreciation per US household ($/yr)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="depr $M / households (M) = $/household/yr").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{annual_depr}/{aref('us_households_m', scen)},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '$#,##0.00'
        cell.fill = TOTAL_FILL
    r += 2

    write_section_label(ws, r, "View B: Industry-wide footprint (5-year committed spend)"); r += 1

    ws.cell(row=r, column=1, value="(Anthropic + OpenAI + one baseline-lab proxy; not a literal industry total)").font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1

    ws.cell(row=r, column=1, value="Industry 5-year commitment ($B)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Implicit Demand'!{cl}{idem['total_commit']}")
        cell.font = GREEN
        cell.number_format = '$#,##0'
    ind_commit = r; r += 1

    ws.cell(row=r, column=1, value="Industry annual run rate ($B/yr)").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"='Implicit Demand'!{cl}{idem['total_annual']}")
        cell.font = GREEN
        cell.number_format = '$#,##0.0'
    ind_annual = r; r += 2

    ws.cell(row=r, column=1, value="Annual rate as % of US GDP").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="annual $B / (GDP $T x 1000)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{ind_annual}/({aref('us_gdp_t', scen)}*1000),0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.00%'
        cell.fill = TOTAL_FILL
    r += 1

    ws.cell(row=r, column=1, value="Annual rate as % of US business capex").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="annual $B / (business capex $T x 1000)").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{ind_annual}/({aref('us_business_capex_t', scen)}*1000),0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%'
        cell.fill = TOTAL_FILL
    r += 2

    write_section_label(ws, r, "Implied physical buildout"); r += 1

    ws.cell(row=r, column=1, value="Hardware share of committed spend").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cell = ws.cell(row=r, column=col, value=f"={aref('capex_share_of_spend', scen)}")
        cell.font = GREEN
        cell.number_format = '0.0%'
    hw_share = r; r += 1

    ws.cell(row=r, column=1, value="Implied industry hardware capex ($B/yr)").font = BLACK
    ws.cell(row=r, column=2, value="annual rate x hardware share").font = BLACK
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"={cl}{ind_annual}*{cl}{hw_share}")
        cell.font = BLACK
        cell.number_format = '$#,##0.0'
    hw_capex = r; r += 1

    ws.cell(row=r, column=1, value="Implied annual DC MW added").font = BLACK_BOLD
    ws.cell(row=r, column=2, value="hardware $B x 1000 / DC cost per MW").font = BLACK
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{hw_capex}*1000/{aref('dc_cost_per_mw', scen)},0)")
        cell.font = BLACK_BOLD
        cell.number_format = '#,##0'
        cell.fill = TOTAL_FILL
    new_mw = r; r += 1

    ws.cell(row=r, column=1, value="New DC MW as % of US grid peak").font = BLACK_BOLD
    for col, scen in zip([3, 4, 5], ['C', 'D', 'E']):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col,
                       value=f"=IFERROR({cl}{new_mw}/({aref('us_grid_peak_gw', scen)}*1000),0)")
        cell.font = BLACK_BOLD
        cell.number_format = '0.0%'
        cell.fill = TOTAL_FILL
    r += 2

    r = write_takeaway_block(ws, r, [
        "Per-lab capex of $30-50B is bps-level vs US GDP but a single-digit % of total US business investment.",
        "Industry annual run rate at Mid (~$120B/yr) is roughly 0.4% of US GDP - measurable but not implausible at the macro level.",
        "Physical constraint is grid capacity: implied new DC builds approach single-digit % of US peak demand annually.",
        "The macro test isn't 'can the economy afford it' - it's 'can power/grid/permitting absorb it on this schedule.'",
    ], row_height=28)

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 38
    for c in ['C', 'D', 'E']:
        ws.column_dimensions[c].width = 14


# ============================================================================
# MAIN
# ============================================================================

def build_model(output_path):
    wb = Workbook()

    build_readme(wb)
    rows = build_assumptions(wb)
    cb = build_cost_build(wb, rows)
    rv = build_revenue_build(wb, rows)
    build_unit_economics(wb, rows, cb, rv)
    build_sensitivity(wb, cb)
    build_build_vs_rent(wb, rows, cb)
    build_breakeven(wb, rows, cb, rv)
    build_no_subsidy(wb, rows, cb)
    idem = build_implicit_demand(wb, rows, cb)
    build_plausibility(wb, rows, idem)
    build_capex_perspective(wb, rows, cb, idem)
    pre = build_stacks_pre(wb, cb, rv)
    build_stacks_post(wb, cb, rv, pre)

    wb.save(output_path)
    print(f"Saved AI True Cost Model to: {output_path}")


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "model.xlsx"
    build_model(output)
